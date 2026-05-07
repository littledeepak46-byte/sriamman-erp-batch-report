"""
M1.25 Batch Slip PDF Generator
===============================
Fills New_Batching_Slip.xlsx with delivery data,
shows/hides batch rows via Column X flags,
converts to PDF via LibreOffice headless.

Place New_Batching_Slip.xlsx in:  backend/templates/New_Batching_Slip.xlsx
"""
import io
import math
import os
import shutil
import subprocess
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties, WorksheetProperties
from pypdf import PdfReader, PdfWriter

TEMPLATE_PATH = Path(__file__).parent.parent.parent / "templates" / "New_Batching_Slip.xlsx"
SHEET_NAME    = "M1.25"


# ── Cell map ──────────────────────────────────────────────────────────────────
# Maps field name → Excel cell address in the M1.25 sheet.
# !! UPDATE THESE once you share / inspect New_Batching_Slip.xlsx !!
CELL_MAP = {
    "batch_date":       None,   # e.g. "C3"
    "batch_start_time": None,   # e.g. "C4"
    "batch_end_time":   None,
    "batch_number":     None,
    "batcher_name":     None,
    "order_number":     None,
    "customer":         None,
    "site":             None,
    "recipe_code":      None,
    "recipe_name":      None,
    "truck_number":     None,
    "truck_driver":     None,
    "ordered_qty":      None,
    "production_qty":   None,
    "adj_manual_qty":   None,
    "with_this_load":   None,
    "batch_size":       None,
    "mixer_capacity":   None,
}

# The Excel row where the FIRST Column-X flag ('show'/'hide') appears.
# Each flag controls 5 rows (Target, Actual, Correction1, Correction2, Separator).
# !! UPDATE after inspecting the template !!
FIRST_FLAG_ROW = None   # e.g. 25


def _set(ws, cell_addr, value):
    """Set a cell value if the address is configured."""
    if cell_addr:
        ws[cell_addr] = value


def _fill_header(ws, data: dict):
    """Fill all header / info cells from the cell map."""
    _set(ws, CELL_MAP["batch_date"],       data.get("batch_date",       ""))
    _set(ws, CELL_MAP["batch_start_time"], data.get("batch_start_time", ""))
    _set(ws, CELL_MAP["batch_end_time"],   data.get("batch_end_time",   ""))
    _set(ws, CELL_MAP["batch_number"],     data.get("batch_number",     ""))
    _set(ws, CELL_MAP["batcher_name"],     "Stetter")
    _set(ws, CELL_MAP["order_number"],     data.get("order_number",     ""))
    _set(ws, CELL_MAP["customer"],         data.get("customer",         ""))
    _set(ws, CELL_MAP["site"],             data.get("site",             ""))
    _set(ws, CELL_MAP["recipe_code"],      data.get("recipe_code",      ""))
    _set(ws, CELL_MAP["recipe_name"],      data.get("recipe_name",      ""))
    _set(ws, CELL_MAP["truck_number"],     data.get("truck_number",     ""))
    _set(ws, CELL_MAP["truck_driver"],     data.get("truck_driver",     ""))
    _set(ws, CELL_MAP["ordered_qty"],      data.get("ordered_qty",      0))
    _set(ws, CELL_MAP["production_qty"],   data.get("production_qty",   0))
    _set(ws, CELL_MAP["adj_manual_qty"],   0)
    _set(ws, CELL_MAP["with_this_load"],   data.get("with_this_load",   0))
    _set(ws, CELL_MAP["batch_size"],       data.get("batch_size",       1.25))
    _set(ws, CELL_MAP["mixer_capacity"],   1.25)


def _fill_batch_rows(ws, batch_actuals: list, num_batches: int):
    """
    Show/hide batch row groups based on Column X flags.
    Each flag in column 24 (X) controls 5 consecutive rows.

    If FIRST_FLAG_ROW is not configured, only the show/hide logic
    from the instructions is applied (scanning all of column 24).
    """
    # Strategy A: scan all of column 24 for show/hide flags (matches instructions)
    flags_found = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 24 and cell.value in ("show", "hide"):
                # show the first num_batches groups, hide the rest
                should_show = flags_found < num_batches
                for sub in range(5):
                    ws.row_dimensions[cell.row + sub].hidden = not should_show
                flags_found += 1

    # Strategy B: use FIRST_FLAG_ROW if configured
    if FIRST_FLAG_ROW and flags_found == 0:
        MAX_BATCH_SLOTS = 20  # maximum slots in the template
        for slot in range(MAX_BATCH_SLOTS):
            flag_row = FIRST_FLAG_ROW + slot * 5
            should_show = slot < num_batches
            for sub in range(5):
                try:
                    ws.row_dimensions[flag_row + sub].hidden = not should_show
                except Exception:
                    break


def generate_m125_pdf(delivery_data: dict, batch_actuals: list) -> bytes:
    """
    Main entry point.

    delivery_data keys:
        batch_date, batch_start_time, batch_number, customer, site,
        recipe_code, truck_number, truck_driver,
        ordered_qty, production_qty, with_this_load, batch_size, num_batches
    batch_actuals: list of per-batch actual value dicts

    Returns: raw PDF bytes
    Raises: FileNotFoundError if template is missing
            RuntimeError if LibreOffice fails
    """
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Excel template not found.\n"
            f"Place New_Batching_Slip.xlsx at:\n  {TEMPLATE_PATH}"
        )

    num_batches = int(delivery_data.get("num_batches", 1))

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_xlsx = os.path.join(tmpdir, "batch_output.xlsx")
        shutil.copy(str(TEMPLATE_PATH), tmp_xlsx)

        wb = openpyxl.load_workbook(tmp_xlsx)

        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet '{SHEET_NAME}' not found in Excel template. "
                             f"Available: {wb.sheetnames}")

        ws = wb[SHEET_NAME]

        # ── Fill header cells ─────────────────────────────────────────────────
        _fill_header(ws, delivery_data)

        # ── Show/hide batch rows ──────────────────────────────────────────────
        _fill_batch_rows(ws, batch_actuals, num_batches)

        # ── Hide all sheets except M1.25 ──────────────────────────────────────
        for sheet_name in wb.sheetnames:
            if sheet_name != SHEET_NAME:
                wb[sheet_name].sheet_state = "hidden"
        wb.active = wb[SHEET_NAME]

        # ── Page setup — Portrait A4, fit to width ────────────────────────────
        ws.sheet_properties = WorksheetProperties()
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation    = "portrait"
        ws.page_setup.paperSize      = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth     = 1
        ws.page_setup.fitToHeight    = 0
        ws.page_margins              = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

        wb.save(tmp_xlsx)

        # ── Convert to PDF via LibreOffice headless ───────────────────────────
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", tmpdir, tmp_xlsx],
            timeout=90,
            capture_output=True,
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"LibreOffice PDF conversion failed:\n"
                f"stdout: {result.stdout.decode()}\n"
                f"stderr: {result.stderr.decode()}"
            )

        pdf_path = os.path.join(tmpdir, "batch_output.pdf")
        if not os.path.exists(pdf_path):
            raise FileNotFoundError("LibreOffice did not produce a PDF output file.")

        # ── Extract M1.25 page (page containing 'Batch Date' + 'Total Set Weight')
        reader = PdfReader(pdf_path)
        writer = PdfWriter()

        for page in reader.pages:
            text = page.extract_text() or ""
            if "Batch Date" in text and "Total Set Weight" in text:
                writer.add_page(page)
                break

        if len(writer.pages) == 0:
            # Fallback: return first page if marker text not found
            if reader.pages:
                writer.add_page(reader.pages[0])

        out = io.BytesIO()
        writer.write(out)
        return out.getvalue()
